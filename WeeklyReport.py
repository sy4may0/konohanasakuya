#coding: UTF-8 
import datetime
import re
import openpyxl
from openpyxl.styles.borders import Border, Side
from openpyxl.styles.alignment import Alignment 
from openpyxl.styles import PatternFill, Font
from collections import OrderedDict


DATE_FORMAT = "%Y-%m-%d"
VISIBLE_DATE_FORMAT = "%Y{0}%m{1}%d{2}"
WEEK_TABLE = ["月","火","水","木","金","土","日"] 

ALIGN_CENTER = Alignment(
    horizontal='center',
    vertical='center',
)

ALIGN_LEFT_WRAP = Alignment(
    horizontal='left',
    vertical='center',
    wrapText=True
)

COLOR_ACCENT = PatternFill(
    patternType='solid',
    fgColor='ffffcc'
)
COLOR_STRONG_ACCENT = PatternFill(
    patternType='solid',
    fgColor='ffc000'
)
COLOR_HOLIDAY = PatternFill(
    patternType='solid',
    fgColor='fde9d9'
)

FONT_BOLD = Font(name='ＭＳ Ｐゴシック', b=True)

class WeeklyReport():
    # achievements ->
    # [{
    #   date: "yyyy-mm-dd"
    #   unexpected: {true|false}
    #   project: xxx 
    #   category: xxx 
    #   description: xxx 
    #   scheduled: xxx 
    #   actual: xxx 
    #   closed: {true|false} 
    #   issues : xxx
    # }...]
    def __init__(self, user, achievements):
        self.__user = user
        self.__achievements = achievements
        self.__excel = {
            "data": OrderedDict(),
            "color": {},
            "border": {}
        }
 
    def addAchievement(self, achievement):
        self.__achievements.append(achievement)

    def extendAchievements(self, achievements):
        self.__achievements.extend(achievements)

    def getAchievements(self):
        return self.__achievements
    
    def searchAchievements(self, filters):
        result = []
        for achievement in self.__achievements:
            checkList = []
            for k in filters:
                if k not in achievement:
                    continue
                if achievement[k] == filters[k]:
                    checkList.append(True)
                else:
                    checkList.append(False)
            if False not in checkList:
                result.append(achievement)
        return result

    # AchievementをExcelのカレンダーデータに追加する。
    # 月曜日をキーとしたカレンダーデータを生成しAchievementを追加する。
    def __addAchievement2Cal(self, achievement):
        # 月曜日算出
        strRefDate = achievement['date']
        refDate = datetime.datetime.strptime(strRefDate,DATE_FORMAT)
        refMonday = refDate - datetime.timedelta(days=refDate.weekday())
        strRefMonday = refMonday.strftime(DATE_FORMAT)

        # Calendar生成済みの場合そこに追加してreturn.
        if strRefMonday in self.__excel["data"]:
            achieveList = self.__excel['data'][strRefMonday][strRefDate]
            achieveList.append(achievement)
            return True

        data = self.__excel["data"] 
        data[strRefMonday] = OrderedDict()

        # 週報の頭にある%Y年%m月%d日(%M)～%Y年%m月%d日(%M)みたいなやつ。
        descr = []
        # formatはWindows対策。strftime直で日本語入れるとエラーになる。
        descr.append(refMonday.strftime(VISIBLE_DATE_FORMAT)
                            .format(*'年月日'))
        descr.append('(' + WEEK_TABLE[refMonday.weekday()] + ')～')

        for i in range(7):
            d = refMonday + datetime.timedelta(days=i)
            strD = d.strftime(DATE_FORMAT)
            data[strRefMonday][strD] = [] 

            # 金曜日
            if i == 4:
                # formatはWindows対策。strftime直で日本語入れるとエラーになる。
                descr.append(d.strftime(VISIBLE_DATE_FORMAT)
                            .format(*'年月日'))
                descr.append('(' + WEEK_TABLE[d.weekday()] + ')')

        data[strRefMonday]['descr'] = ''.join(descr)
        data[strRefMonday]['sheetname'] = strRefMonday.replace('-','') 

        achieveList = self.__excel['data'][strRefMonday][strRefDate]
        achieveList.append(achievement)
       
        return True 

    # ExcelのborderテーブルにBorderを追加する。
    def __addBorder(self, style, cols, rows, edges):
        borderTable = self.__excel['border']
        side = Side(style= style ,color='000000')
        for col in cols:
            for row in rows:
                target = col+str(row)
                if target not in borderTable:
                    borderTable[target]={
                        "top":Side(style=None), 
                        "left":Side(style=None), 
                        "right":Side(style=None), 
                        "bottom":Side(style=None)}
                for edge in edges:
                    borderTable[target][edge]=side

    # カレンダーデータから1週間分取り出してExcelのシートを作成。
    def __addWeekBlock(self, excelBook, weekData):
        excelBook.create_sheet(index=0,title=weekData['sheetname'])
        excelSheet = excelBook[weekData['sheetname']]

        # Set Column dimensions
        colDimensionsTable = {
            "A": 3.0,
            "B": 11.0,
            "C": 11.0,
            "D": 18.5,
            "E": 28.5,
            "F": 62.0,
            "G": 6.0,
            "H": 6.0,
            "I": 11.0,
            "J": 62.0,
        }
        for k in colDimensionsTable:
            excelSheet.column_dimensions[k].width \
                    = colDimensionsTable[k]
        # Set Row dimensions
        rowDimensionsTable = {
            1: 7.5,
            2: 18.0,
            3: 18.0,
            5: 15.0
        }
        for k in rowDimensionsTable:
            excelSheet.row_dimensions[k].height \
                    = rowDimensionsTable[k]

        # Create Username And Title Section
        # [data]
        excelSheet['B2']="氏名"
        excelSheet['B2'].fill = COLOR_ACCENT
        excelSheet['B2'].font = FONT_BOLD
        excelSheet['B3']="報告期間"
        excelSheet['B3'].fill = COLOR_ACCENT
        excelSheet['B3'].font = FONT_BOLD
        excelSheet['C2']=self.__user
        excelSheet['C2'].font = FONT_BOLD
        excelSheet['C3']=weekData['descr']
        # [Merge Cell]
        excelSheet.merge_cells('C2:E2')
        excelSheet.merge_cells('C3:E3')
        # [border]
        cols = list("BCDE")
        rows = [2, 3]
        edges = ["left","right","top","bottom"]
        self.__addBorder('thin', cols, rows, edges)

        # Create Header.
        # [data]
        headers = [
            "作業日",
            "予定外作業",
            "システム名",
            "作業種別",
            "作業内容",
            "予定",
            "実績",
            "ステータス",
            "課題/問題点など",
        ]
        col = 2
        for h in headers:
            excelSheet.cell(row=5, column=col).value = h 
            excelSheet.cell(row=5, column=col).alignment = ALIGN_CENTER
            excelSheet.cell(row=5, column=col).fill = COLOR_ACCENT
            excelSheet.cell(row=5, column=col).font = FONT_BOLD
            col+=1
        # [border]
        cols = list("BCDEFGHIJ") 
        rows = [5]
        edges = ["left","right","top","bottom"]
        self.__addBorder('thin', cols, rows, edges)

        dayBlockStart = 6
        blockIndex = 0
        for dateKey in weekData.keys():
            if re.match('[0-9]{4}-[0-9]{2}-[0-9]{2}', dateKey):
                holiday = False
                # 日曜日とどようびどはholiday=True
                if blockIndex >= 5:
                    holiday = True
                dayBlockStart = self.__addDayBlock(
                                    excelSheet, 
                                    dayBlockStart, 
                                    dateKey,
                                    weekData[dateKey],
                                    holiday)
                blockIndex+=1

        # Activate Border lines.
        for k in self.__excel['border']:
            side = self.__excel['border'][k]
            border = Border(
                top=side['top'],
                left=side['left'],
                right=side['right'],
                bottom=side['bottom']
            )
            excelSheet[k].border = border
        self.__excel['border'] = {}

    # Scheduled/Actualの合計を算出。
    # 小計用。
    def __getSubTotal(self, dayBlocks):
        totalScheduled = {"hour": 0, "minute": 0}
        totalActual = {"hour": 0, "minute": 0}
        for dayBlock in dayBlocks:

            if re.match('[0-9]+:[0-9]+',dayBlock['scheduled']):
                splitTimeMin = dayBlock['scheduled'].split(':')
                hour = int(splitTimeMin[0])
                minute = int(splitTimeMin[1])
                totalScheduled['hour']+= hour
                totalScheduled['minute']+= minute

            if re.match('[0-9]+:[0-9]+',dayBlock['actual']):
                splitTimeMin = dayBlock['actual'].split(':')
                hour = int(splitTimeMin[0])
                minute = int(splitTimeMin[1])
                totalActual['hour']+=hour
                totalActual['minute']+=minute

        totalScheduled['hour'] += totalScheduled['minute']//60
        totalScheduled['minute'] = totalScheduled['minute']%60
        totalActual['hour'] += totalActual['minute']//60
        totalActual['minute'] = totalActual['minute']%60

        scheduled = str(totalScheduled['hour']) + ':' + \
                    str(totalScheduled['minute']).zfill(2)
        actual = str(totalActual['hour']) + ':' + \
                 str(totalActual['minute']).zfill(2)
        subTotal = {
            "scheduled": scheduled,
            "actual": actual 
        }

        return subTotal
        
    # 1日分のAchievementをExcelに追記。
    def __addDayBlock(self, excelSheet, startRow, day, dayBlocks, holiday=False):
        blockLength = len(dayBlocks) + 2
        endRow = startRow + blockLength

        # Create Day Block head.
        # [data]
        excelSheet.cell(row=startRow, column=2).value = day.replace('-','/')
        excelSheet.cell(row=startRow, column=2).alignment = ALIGN_CENTER
        if holiday:
            excelSheet.cell(row=startRow, column=2).fill = COLOR_HOLIDAY

        # [Merge cell]
        mergeRange = 'B'+ str(startRow) + ':B' + str(endRow)
        excelSheet.merge_cells(mergeRange)
        # [border]
        # 囲む。
        cols = list("B")
        self.__addBorder('thin', cols, [startRow], ['top'])
        rows = list(range(startRow, endRow+1)) 
        edges = ["left","right"]
        self.__addBorder('thin', cols, rows, edges)
        self.__addBorder('thin', cols, [endRow], ['bottom'])

        # Create Day Block
        # [data]
        dayBlockRow = startRow
        for dayBlock in dayBlocks:
            attr = [
                '',
                dayBlock['project'],
                dayBlock['category'],
                dayBlock['description'],
                dayBlock['scheduled'],
                dayBlock['actual'],
                'OPEN',
                dayBlock['issues']
            ]
            if dayBlock['unexpected']:
                attr[0] = "予定外"
            if dayBlock['closed']:
                attr[6] = "CLOSE"

            for i in range(3,11):
                excelSheet.cell(row=dayBlockRow, column=i).value = \
                               attr[i-3]
                excelSheet.cell(row=dayBlockRow, column=i).alignment = \
                                ALIGN_LEFT_WRAP
            # GHI列はALIGN_CENTER
            for i in range(7,10):
                excelSheet.cell(row=dayBlockRow, column=i).alignment = \
                                ALIGN_CENTER
            dayBlockRow+=1

        # [border]
        # 囲む。かつ中に点線。
        cols = list("CDEFGHIJ")
        self.__addBorder('thin', cols, [startRow], ['top'])
        rows = list(range(startRow, endRow+1)) 
        edges = ["left","right"]
        self.__addBorder('thin', cols, rows, edges)
        edges = ["bottom"]
        self.__addBorder('dashed', cols, rows, edges)
        self.__addBorder('thin', cols, [endRow], ['bottom'])

        # Create Day Block subTotal
        # [data]
        subTotal = self.__getSubTotal(dayBlocks)
        excelSheet.cell(row=endRow, column=6).value = "小計" 
        excelSheet.cell(row=endRow, column=6).alignment = ALIGN_CENTER
        excelSheet.cell(row=endRow, column=6).font = FONT_BOLD
        excelSheet.cell(row=endRow, column=7).value = subTotal['scheduled'] 
        excelSheet.cell(row=endRow, column=7).alignment = ALIGN_CENTER
        excelSheet.cell(row=endRow, column=8).value = subTotal['actual'] 
        excelSheet.cell(row=endRow, column=8).alignment = ALIGN_CENTER
        # [background color]
        for col in list(range(3,11)):
            excelSheet.cell(row=endRow, column=col).fill \
                             = COLOR_STRONG_ACCENT 
        return endRow + 1 


    def writeToExcel(self, resultPath):
        excelBook = openpyxl.Workbook(write_only=False)

        for achievement in self.__achievements:
            self.__addAchievement2Cal(achievement)

        for k in self.__excel['data'].keys():
            self.__addWeekBlock(excelBook, self.__excel['data'][k])

        excelBook.save(resultPath)


                
